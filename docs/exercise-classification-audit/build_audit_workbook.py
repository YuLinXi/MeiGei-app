#!/usr/bin/env python3
import csv
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
ALL_CSV = ROOT / "all-exercises.csv"
FLAGGED_CSV = ROOT / "flagged-exercises.csv"
OUTPUT_XLSX = ROOT / "动作分类审核表.xlsx"

REGION_NAMES = {
    "traps": "斜方肌",
    "deltFront": "三角肌前束",
    "deltRear": "三角肌后束",
    "chest": "胸大肌",
    "biceps": "肱二头肌",
    "triceps": "肱三头肌",
    "forearms": "前臂",
    "abs": "腹直肌",
    "obliques": "腹外斜肌",
    "lats": "背阔肌",
    "lowerBack": "竖脊肌",
    "glutes": "臀大肌",
    "quads": "股四头肌",
    "adductors": "内收肌",
    "hams": "腘绳肌",
    "calves": "小腿",
    "deltSide": "三角肌中束",
    "rhomboids": "菱形肌",
    "gluteMed": "臀中肌",
    "neck": "颈部",
}

OPTION_GROUPS = {
    "审核结论": ["通过", "调整分类", "删除", "合并", "改名", "待确认"],
    "一级分类": ["胸", "背", "肩", "手臂", "腿", "臀", "核心", "颈部", "有氧", "功能性", "热身拉伸"],
    "二级肌肉": [
        "胸大肌", "背阔肌", "斜方肌", "菱形肌", "竖脊肌", "三角肌", "肱二头肌", "肱三头肌", "前臂",
        "股四头肌", "腘绳肌", "内收肌", "小腿", "臀大肌", "臀中肌", "腹直肌", "腹斜肌", "颈部肌",
    ],
    "三级肌区": [
        "上胸", "中下胸", "前束", "中束", "后束", "上腹", "下腹",
        "爆发奥举", "抗旋稳定", "髋臀激活", "负重搬运", "动态控制",
        "动态热身", "静态拉伸", "泡沫轴放松",
    ],
    "器械": ["杠铃", "哑铃", "壶铃", "器械", "史密斯", "悍马机", "T杠", "绳索", "弹力带", "悬挂", "自重", "其他"],
    "肌区": list(REGION_NAMES.values()),
}

REVIEW_HEADERS = [
    "序号",
    "来源",
    "动作编码",
    "动作名称",
    "原始一级分类",
    "原始子分类",
    "器械",
    "实际一级分类",
    "实际二级肌肉",
    "实际三级肌区",
    "主动肌区",
    "协同肌区",
    "用户可见路径",
    "可疑提示",
    "审核结论",
    "审核备注",
]

EDITABLE_HEADERS = {
    "器械",
    "实际一级分类",
    "实际二级肌肉",
    "实际三级肌区",
    "主动肌区",
    "协同肌区",
    "审核结论",
    "审核备注",
}


def read_csv(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.reader(file))


def to_chinese_regions(value: str) -> str:
    if not value:
        return ""
    parts = [part for part in value.split("|") if part]
    return "、".join(REGION_NAMES.get(part, part) for part in parts)


def localized_rows(path: Path) -> list[list[str]]:
    rows = read_csv(path)
    if not rows:
        return rows
    header = rows[0]
    active_idx = header.index("主动肌区")
    secondary_idx = header.index("协同肌区")
    output = [REVIEW_HEADERS]
    for row in rows[1:]:
        row = row[:]
        row[active_idx] = to_chinese_regions(row[active_idx])
        row[secondary_idx] = to_chinese_regions(row[secondary_idx])
        by_header = dict(zip(header, row))
        output.append([by_header.get(name, "") for name in REVIEW_HEADERS])
    return output


def write_rows(ws, rows: list[list[str]]) -> None:
    for row in rows:
        ws.append(row)


def setup_sheet(ws, table_name: str) -> None:
    max_row = ws.max_row
    max_col = ws.max_column
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    soft_border = Side(style="thin", color="D7D0C3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=soft_border)

    widths = {
        "A": 8, "B": 10, "C": 28, "D": 28,
        "E": 14, "F": 16, "G": 12, "H": 14,
        "I": 16, "J": 16, "K": 24, "L": 30,
        "M": 26, "N": 22, "O": 14, "P": 36,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 28

    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="E8E1D6"))

    editable_fill = PatternFill("solid", fgColor="FFF4CC")
    headers = [cell.value for cell in ws[1]]
    for col_idx, header in enumerate(headers, start=1):
        if header in EDITABLE_HEADERS:
            for row_idx in range(2, max_row + 1):
                ws.cell(row=row_idx, column=col_idx).fill = editable_fill

    ref = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
    ws.add_table(table)


def write_options_sheet(wb: Workbook):
    ws = wb.create_sheet("选项")
    ws.append(list(OPTION_GROUPS.keys()))
    max_len = max(len(values) for values in OPTION_GROUPS.values())
    for row_idx in range(max_len):
        ws.append([
            values[row_idx] if row_idx < len(values) else ""
            for values in OPTION_GROUPS.values()
        ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EDE6D8")
    for col_idx in range(1, len(OPTION_GROUPS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18
    ws.sheet_state = "hidden"
    return ws


def option_range(group_name: str) -> str:
    group_names = list(OPTION_GROUPS.keys())
    col_idx = group_names.index(group_name) + 1
    col_letter = chr(ord("A") + col_idx - 1)
    end_row = len(OPTION_GROUPS[group_name]) + 1
    return f"'选项'!${col_letter}$2:${col_letter}${end_row}"


def add_dropdown(ws, header_name: str, option_group: str, prompt: Optional[str] = None) -> None:
    headers = [cell.value for cell in ws[1]]
    if header_name not in headers:
        return
    col_idx = headers.index(header_name) + 1
    col_letter = ws.cell(row=1, column=col_idx).column_letter
    dv = DataValidation(type="list", formula1=option_range(option_group), allow_blank=True)
    dv.showErrorMessage = False
    dv.showInputMessage = bool(prompt)
    if prompt:
        dv.promptTitle = "填写提示"
        dv.prompt = prompt
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{max(ws.max_row, 2)}")


def add_dropdowns(ws) -> None:
    add_dropdown(ws, "主动肌区", "肌区", "下拉为单个肌区；如需多个肌区，可手动用顿号分隔。")
    add_dropdown(ws, "协同肌区", "肌区", "下拉为单个肌区；如需多个肌区，可手动用顿号分隔。")
    add_dropdown(ws, "审核结论", "审核结论")
    add_dropdown(ws, "实际一级分类", "一级分类")
    add_dropdown(ws, "实际二级肌肉", "二级肌肉")
    add_dropdown(ws, "实际三级肌区", "三级肌区")
    add_dropdown(ws, "器械", "器械")


def write_instruction_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("说明", 0)
    rows = [
        ["动作分类审核表"],
        ["使用方式"],
        ["1. 优先看「可疑优先」工作表，处理脚本标出的明显错位。"],
        ["2. 黄色列可直接修改：器械、实际一级分类、实际二级肌肉、实际三级肌区、主动肌区、协同肌区、审核结论、审核备注。"],
        ["3. 「主动肌区」「协同肌区」已经改成中文显示，并带中文下拉选项。"],
        ["4. Excel 原生下拉一次只能选一个肌区；多个肌区请手动用顿号分隔，例如：三角肌前束、肱三头肌。"],
        ["5. 审核结论建议填：通过 / 调整分类 / 删除 / 合并 / 改名 / 待确认。"],
        ["6. 审完后只需要保留/告诉我「审核结论」和「审核备注」，我会读取你直接调整后的原列并回填到 Swift 动作数据。"],
    ]
    for row in rows:
        ws.append(row)
    ws["A1"].font = Font(size=18, bold=True, color="7A2E1F")
    ws["A2"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 96
    for cell in ws["A"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.sheet_view.showGridLines = False


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    write_instruction_sheet(wb)
    write_options_sheet(wb)

    all_ws = wb.create_sheet("全量审核")
    write_rows(all_ws, localized_rows(ALL_CSV))
    setup_sheet(all_ws, "AllExercises")
    add_dropdowns(all_ws)

    flagged_ws = wb.create_sheet("可疑优先")
    write_rows(flagged_ws, localized_rows(FLAGGED_CSV))
    setup_sheet(flagged_ws, "FlaggedExercises")
    add_dropdowns(flagged_ws)

    wb.active = 0
    wb.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
